"""Round-trip tests for CaixaBank .xlsx parsing (Excel serial dates, header variants)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from src.parsers.caixabank import _description_from_cc, _parse_date_cell, parse_caixabank_file
from src.ingest import load_and_parse_source
from src.config_manager import load_config


# ---------------------------------------------------------------------------
# Unit tests
# ---------------------------------------------------------------------------

def test_parse_date_cell_excel_serial() -> None:
    serial = to_excel(datetime(2026, 4, 3))
    d = _parse_date_cell(None, serial)
    assert d is not None and d.isoformat() == "2026-04-03"


def test_parse_date_cell_datetime_object() -> None:
    d = _parse_date_cell(None, datetime(2026, 4, 3, 12, 0))
    assert d is not None and d.isoformat() == "2026-04-03"


def test_description_from_cc_skips_empty() -> None:
    cells = {"cc1": "MERCHANT A", "cc2": "", "cc3": "ES400", "cc5": "  NOTE  "}
    for k in range(1, 11):
        cells.setdefault(f"cc{k}", "")
    assert _description_from_cc(cells) == "MERCHANT A ES400 NOTE"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_wide_xlsx(path: Path, *, valor_cell: Any, oper_cell: Any) -> None:  # type: ignore[name-defined]
    """Write a minimal CaixaBank wide-format .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Oficina", "Divisa", "F. Operación", "F. Valor",
        "Ingreso (+)", "Gasto (-)", "Saldo (+)", "Saldo (-)",
        "Concepto común", "Concepto propio", "Referencia 1", "Referencia 2",
    ] + [f"Concepto complementario {k}" for k in range(1, 11)])
    ws.append([
        "0599", "EUR", oper_cell, valor_cell,
        0, 111, 5147.11, "",
        "03", "034", "000000000", "080569617812",
        "MERCHANT A", "", "ES40000B60637766", "",
        "LICENCIA FED", "", "", "",
        "EXAMPLE MERCHANT B", "",
    ])
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Integration tests
# ---------------------------------------------------------------------------

def test_xlsx_parses_serial_dates(tmp_path: Path) -> None:
    p = tmp_path / "caixa.xlsx"
    serial = to_excel(datetime(2026, 4, 3))
    _make_wide_xlsx(p, valor_cell=serial, oper_cell=serial)
    rows = parse_caixabank_file(p, "test")
    assert len(rows) == 1
    r = rows[0]
    assert r["date"] == "2026-04-03"
    assert r["amount"] == -111.0
    assert "MERCHANT A" in r["description"]
    assert "EXAMPLE MERCHANT B" in r["description"]
    assert r["cb_gasto"] == 111.0


def test_xlsx_parses_datetime_cells(tmp_path: Path) -> None:
    p = tmp_path / "caixa2.xlsx"
    _make_wide_xlsx(p, valor_cell=datetime(2026, 4, 3), oper_cell=datetime(2026, 4, 3))
    rows = parse_caixabank_file(p, "test")
    assert len(rows) == 1
    assert rows[0]["date"] == "2026-04-03"


def test_xlsx_header_without_space_f_valor(tmp_path: Path) -> None:
    """Bank sometimes exports 'F.Valor' without the space."""
    p = tmp_path / "caixa3.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["F.Valor", "Ingreso (+)", "Gasto (-)", "Concepto complementario 1"])
    ws.append([to_excel(datetime(2025, 12, 1)), 0, 50, "SHOP"])
    wb.save(p)
    wb.close()
    rows = parse_caixabank_file(p, "test")
    assert len(rows) == 1
    assert rows[0]["amount"] == -50.0
    assert "SHOP" in rows[0]["description"]


def test_xlsx_blank_leading_column(tmp_path: Path) -> None:
    """File has a blank column A (col 0) — must not confuse column mapping."""
    p = tmp_path / "caixa4.xlsx"
    wb = Workbook()
    ws = wb.active
    # col 0 is blank, data starts at col 1
    ws.append(["", "F. Valor", "Ingreso (+)", "Gasto (-)", "Concepto complementario 1"])
    ws.append(["", datetime(2026, 1, 15), 0, 75.5, "SUPERMARKET"])
    wb.save(p)
    wb.close()
    rows = parse_caixabank_file(p, "test")
    assert len(rows) == 1
    assert rows[0]["amount"] == -75.5
    assert "SUPERMARKET" in rows[0]["description"]


def test_full_ingest_pipeline_caixabank_joint() -> None:
    """Full load_and_parse_source path; skipped if the local file is absent."""
    from pathlib import Path as P
    import pytest
    root = P(__file__).resolve().parents[1]
    xls = root / "tmp" / "input" / "caixabank_joint.XLS"
    if not xls.is_file():
        xls = root / "tmp" / "input" / "caixabank_joint.xls"
    if not xls.is_file():
        pytest.skip("missing tmp/input/caixabank_joint.xls")

    cfg = load_config()
    spec = next(s for s in cfg["bank_imports"]["sources"] if s["id"] == "caixabank_joint")
    rows = load_and_parse_source(cfg, spec)
    assert len(rows) > 0, "Expected >0 rows from caixabank_joint"
    r = rows[0]
    assert r["date"], "row must have a date"
    assert r["amount"] != 0, "row must have a non-zero amount"
    assert r["source"] == "caixabank_joint"
    # description must come only from cc fields
    assert r["description"] is not None
