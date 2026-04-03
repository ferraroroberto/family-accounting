"""Parse CaixaBank Excel exports (.xls via xlrd, .xlsx via openpyxl).

Column layout is auto-detected by scanning for the header row (up to row 50).
The two supported formats are:

Wide export (CaixaBankNow full detail):
  Requires columns: F. Valor (or F. Operación) + Ingreso (+) + Gasto (-)
  • transaction date  = F. Valor  (falls back to F. Operación when missing)
  • value_date        = F. Operación (stored for reference)
  • amount            = Ingreso (+) − Gasto (-)  [European number format supported]
  • description       = non-empty Concepto complementario 1…10, space-joined
  • all raw Caixa fields stored in cb_* columns (cb_cc1…cb_cc10, cb_oficina, …)

Compact export (older Fecha/Movimiento/Importe format):
  description = Movimiento + Más datos
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Protocol

import xlrd
from xlrd.sheet import Sheet as XlrdSheet


# ---------------------------------------------------------------------------
# Cell value helpers
# ---------------------------------------------------------------------------

def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()


def _norm_header(s: str) -> str:
    """Lowercase, ASCII-fold accented vowels."""
    return (
        _cell_str(s).lower()
        .replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u")
    )


def _parse_date_cell(book: Any | None, cell: Any) -> date | None:
    """Convert xlrd serial, openpyxl datetime/serial, or DD/MM/YYYY string to date."""
    if cell == "" or cell is None or isinstance(cell, bool):
        return None
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    if isinstance(cell, (int, float)):
        # xlrd path
        if book is not None:
            try:
                return xlrd.xldate_as_datetime(float(cell), book.datemode).date()
            except (ValueError, xlrd.XLDateError, TypeError):
                pass
        # openpyxl path (xlsx serial without xlrd book)
        try:
            from openpyxl.utils.datetime import from_excel
            result = from_excel(float(cell))
            if isinstance(result, datetime):
                return result.date()
            if isinstance(result, date):
                return result
        except (ValueError, TypeError, OverflowError):
            pass
    s = _cell_str(cell)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _float_cell(v: Any) -> float:
    """Parse a numeric cell; handles European grouping (2.000,00 → 2000.0)."""
    if v == "" or v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("EUR", "").replace("€", "").strip()
    if not s or s in ("-", "—"):
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1].strip()
    # Detect European vs US format
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):      # 1.234,56 → European
            s = s.replace(".", "").replace(",", ".")
        else:                                 # 1,234.56 → US
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2 and parts[1].isdigit():
            s = parts[0].replace(".", "") + "." + parts[1]
        else:
            s = s.replace(",", "")
    try:
        val = float(s)
    except ValueError:
        return 0.0
    return -val if neg else val


def _iso(d: date | None) -> str | None:
    return d.isoformat() if d else None


# ---------------------------------------------------------------------------
# Sheet abstraction (xlrd vs openpyxl)
# ---------------------------------------------------------------------------

class _SheetLike(Protocol):
    @property
    def nrows(self) -> int: ...
    @property
    def ncols(self) -> int: ...
    def cell_value(self, ri: int, ci: int) -> Any: ...


class _XlrdSheet:
    def __init__(self, book: Any, sheet: XlrdSheet) -> None:
        self._book = book
        self._sheet = sheet

    @property
    def nrows(self) -> int:
        return self._sheet.nrows

    @property
    def ncols(self) -> int:
        return self._sheet.ncols

    def cell_value(self, ri: int, ci: int) -> Any:
        return self._sheet.cell_value(ri, ci)


class _OpxSheet:
    def __init__(self, ws: Any) -> None:
        self._ws = ws

    @property
    def nrows(self) -> int:
        return int(self._ws.max_row or 0)

    @property
    def ncols(self) -> int:
        return int(self._ws.max_column or 0)

    def cell_value(self, ri: int, ci: int) -> Any:
        v = self._ws.cell(row=ri + 1, column=ci + 1).value
        return v if v is not None else ""


def _open_sheet(path: Path, sheet_index: int) -> tuple[Any | None, _SheetLike]:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        # read_only=True can leave max_row/max_column unset on some files
        wb = load_workbook(path, read_only=False, data_only=True)
        return wb, _OpxSheet(wb.worksheets[sheet_index])
    book = xlrd.open_workbook(str(path))
    return book, _XlrdSheet(book, book.sheet_by_index(sheet_index))


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _row_headers(sheet: _SheetLike, ri: int) -> list[str]:
    return [_cell_str(sheet.cell_value(ri, ci)) for ci in range(sheet.ncols)]


def _find_col(headers: list[str], label: str) -> int | None:
    """Return column index for *label* (normalized, skips blank headers).

    Match order: (1) exact, (2) label is substring of header, (3) non-empty
    header is substring of label.
    """
    want = _norm_header(label)
    if not want:
        return None
    for ci, h in enumerate(headers):
        hn = _norm_header(h)
        if hn and hn == want:
            return ci
    for ci, h in enumerate(headers):
        hn = _norm_header(h)
        if hn and (want in hn or hn in want):
            return ci
    return None


def _find_header_row(sheet: _SheetLike, max_scan: int = 50) -> tuple[int, dict[str, int]]:
    """Scan rows to find the CaixaBank header row.

    Returns (row_index, colmap) where colmap maps logical keys to column indices.
    Raises ValueError if no valid header is found within max_scan rows.
    """
    for ri in range(min(max_scan, sheet.nrows)):
        headers = _row_headers(sheet, ri)
        texts = [_norm_header(h) for h in headers]

        # --- Wide export (F. Valor / F. Operación + Ingreso + Gasto) ---
        cm: dict[str, int] = {}
        for ci, t in enumerate(texts):
            if not t:
                continue
            tl = t.replace(" ", "")
            if "oficina" in t and "complementario" not in t:
                cm.setdefault("oficina", ci)
            if t in ("divisa",):
                cm.setdefault("divisa", ci)
            if "f.operacion" in tl:
                cm["f_operacion"] = ci
            if "f.valor" in tl:
                cm["f_valor"] = ci
            if "ingreso" in t and "+" in t:
                cm["ingreso"] = ci
            if "gasto" in t and ("-" in t or "\u2212" in t):
                cm["gasto"] = ci
            if "saldo" in t and "+" in t:
                cm["saldo_pos"] = ci
            elif "saldo" in t and ("-" in t or "\u2212" in t):
                cm["saldo_neg"] = ci
            elif "saldo" in t:
                cm.setdefault("saldo", ci)
            if "concepto comun" in t and "complementario" not in t:
                cm["concepto_comun"] = ci
            if "concepto propio" in t and "complementario" not in t:
                cm["concepto_propio"] = ci
            if "referencia 1" in t:
                cm["referencia1"] = ci
            if "referencia 2" in t:
                cm["referencia2"] = ci
            m = re.match(r"^concepto complementario (\d{1,2})$", t)
            if m:
                n = int(m.group(1))
                if 1 <= n <= 20:
                    cm[f"cc{n}"] = ci

        has_date = "f_valor" in cm or "f_operacion" in cm
        if has_date and {"ingreso", "gasto"}.issubset(cm):
            return ri, cm

        # --- Compact export (Fecha + Movimiento + Importe) ---
        cc: dict[str, int] = {}
        for ci, t in enumerate(texts):
            if t == "fecha":
                cc["fecha"] = ci
            elif t == "fecha valor":
                cc["fecha_valor"] = ci
            elif t == "movimiento":
                cc["movimiento"] = ci
            elif "mas datos" in t or "más datos" in t:
                cc["mas_datos"] = ci
            elif t == "importe":
                cc["importe"] = ci
            elif t == "saldo":
                cc["saldo"] = ci
        if {"fecha", "movimiento", "importe"}.issubset(cc):
            return ri, cc

    raise ValueError(
        "CaixaBank header row not found (scanned %d rows). "
        "Wide export needs F. Valor (or F. Operación) + Ingreso (+) + Gasto (-). "
        "Compact export needs Fecha + Movimiento + Importe." % min(max_scan, sheet.nrows)
    )


# ---------------------------------------------------------------------------
# Null skeleton for cb_* columns (Revolut rows leave these NULL)
# ---------------------------------------------------------------------------

_CB_NULL: dict[str, Any] = {
    "cb_oficina": None, "cb_divisa": None,
    "cb_f_operacion": None, "cb_f_valor": None,
    "cb_ingreso": None, "cb_gasto": None,
    "cb_saldo_pos": None, "cb_saldo_neg": None,
    "cb_concepto_comun": None, "cb_concepto_propio": None,
    "cb_referencia1": None, "cb_referencia2": None,
    **{f"cb_cc{i}": None for i in range(1, 11)},
}


# ---------------------------------------------------------------------------
# Row parsers
# ---------------------------------------------------------------------------

def _cell_at(sheet: _SheetLike, ri: int, colmap: dict[str, int], key: str) -> str:
    if key not in colmap:
        return ""
    return _cell_str(sheet.cell_value(ri, colmap[key]))


def _cell_is_empty(sheet: _SheetLike, ri: int, ci: int) -> bool:
    v = sheet.cell_value(ri, ci)
    return v == "" or v is None


def _description_from_cc(cells: dict[str, str]) -> str:
    """Join non-empty Concepto complementario 1…10, collapsed whitespace."""
    parts = [" ".join((cells.get(f"cc{k}") or "").split()) for k in range(1, 11)]
    return " ".join(p for p in parts if p)


def _parse_wide_rows(
    book: Any | None,
    sheet: _SheetLike,
    colmap: dict[str, int],
    start_row: int,
    source_id: str,
    date_key: str,
) -> list[dict[str, Any]]:
    """Parse wide-format CaixaBank rows into canonical transaction dicts."""
    from src.database import transaction_hash

    text_keys = [
        "oficina", "divisa", "concepto_comun", "concepto_propio",
        "referencia1", "referencia2",
        *[f"cc{k}" for k in range(1, 11)],
    ]
    rows: list[dict[str, Any]] = []

    for ri in range(start_row, sheet.nrows):
        d_valor = _parse_date_cell(book, sheet.cell_value(ri, colmap["f_valor"])) if "f_valor" in colmap else None
        d_oper = _parse_date_cell(book, sheet.cell_value(ri, colmap["f_operacion"])) if "f_operacion" in colmap else None

        if date_key == "f_valor":
            op = d_valor or d_oper
            value_for_db = _iso(d_oper)
        else:
            op = d_oper or d_valor
            value_for_db = _iso(d_valor)

        if op is None:
            continue

        cells = {lk: _cell_at(sheet, ri, colmap, lk) for lk in text_keys if lk in colmap}

        ing = _float_cell(sheet.cell_value(ri, colmap["ingreso"]))
        gas = _float_cell(sheet.cell_value(ri, colmap["gasto"]))
        amt = ing - gas

        bal: float | None = None
        if "saldo_pos" in colmap and not _cell_is_empty(sheet, ri, colmap["saldo_pos"]):
            bal = _float_cell(sheet.cell_value(ri, colmap["saldo_pos"]))
        if bal is None and "saldo_neg" in colmap and not _cell_is_empty(sheet, ri, colmap["saldo_neg"]):
            bal = -abs(_float_cell(sheet.cell_value(ri, colmap["saldo_neg"])))
        if bal is None and "saldo" in colmap and not _cell_is_empty(sheet, ri, colmap["saldo"]):
            bal = _float_cell(sheet.cell_value(ri, colmap["saldo"]))

        desc = _description_from_cc(cells)
        ds = op.isoformat()
        cb: dict[str, Any] = dict(_CB_NULL)
        cb.update({
            "cb_oficina": cells.get("oficina") or None,
            "cb_divisa": cells.get("divisa") or None,
            "cb_f_operacion": _iso(d_oper),
            "cb_f_valor": _iso(d_valor),
            "cb_ingreso": ing or None,
            "cb_gasto": gas or None,
            "cb_concepto_comun": cells.get("concepto_comun") or None,
            "cb_concepto_propio": cells.get("concepto_propio") or None,
            "cb_referencia1": cells.get("referencia1") or None,
            "cb_referencia2": cells.get("referencia2") or None,
            **{f"cb_cc{i}": (cells.get(f"cc{i}") or "").strip() or None for i in range(1, 11)},
        })
        if "saldo_pos" in colmap:
            cb["cb_saldo_pos"] = None if _cell_is_empty(sheet, ri, colmap["saldo_pos"]) else _float_cell(sheet.cell_value(ri, colmap["saldo_pos"]))
        if "saldo_neg" in colmap:
            cb["cb_saldo_neg"] = None if _cell_is_empty(sheet, ri, colmap["saldo_neg"]) else _float_cell(sheet.cell_value(ri, colmap["saldo_neg"]))

        rows.append({
            "source": source_id,
            "date": ds,
            "value_date": value_for_db,
            "description": desc,
            "amount": amt,
            "balance": bal,
            "yyyymm": op.strftime("%Y%m"),
            "hash": transaction_hash(source_id, ds, amt, desc),
            "extra_json": None,
            **cb,
        })
    return rows


def _parse_compact_rows(
    book: Any | None,
    sheet: _SheetLike,
    colmap: dict[str, int],
    start_row: int,
    source_id: str,
) -> list[dict[str, Any]]:
    """Parse compact-format (Fecha/Movimiento/Importe) CaixaBank rows."""
    from src.database import transaction_hash

    rows: list[dict[str, Any]] = []
    for ri in range(start_row, sheet.nrows):
        op = _parse_date_cell(book, sheet.cell_value(ri, colmap["fecha"]))
        if op is None:
            continue
        val = _parse_date_cell(book, sheet.cell_value(ri, colmap.get("fecha_valor", colmap["fecha"])))
        mov = _cell_str(sheet.cell_value(ri, colmap["movimiento"]))
        extra = _cell_at(sheet, ri, colmap, "mas_datos")
        desc = " ".join(p for p in (mov, extra) if p).strip()
        amt = _float_cell(sheet.cell_value(ri, colmap["importe"]))
        bal: float | None = None
        if "saldo" in colmap:
            bal = _float_cell(sheet.cell_value(ri, colmap["saldo"])) or None
        ds = op.isoformat()
        rows.append({
            "source": source_id,
            "date": ds,
            "value_date": _iso(val),
            "description": desc,
            "amount": amt,
            "balance": bal,
            "yyyymm": op.strftime("%Y%m"),
            "hash": transaction_hash(source_id, ds, amt, desc),
            "extra_json": None,
            **_CB_NULL,
        })
    return rows


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_caixabank_file(
    path: Path,
    source_id: str,
    layout: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Parse a CaixaBank .xls / .xlsx export into canonical transaction dicts.

    layout keys (all optional):
      sheet_index             int, default 0
      transaction_date_column "f_valor" (default) | "f_operacion"
    """
    layout = layout or {}
    sheet_index = int(layout.get("sheet_index", 0))
    date_key = layout.get("transaction_date_column", "f_valor")
    if date_key not in ("f_valor", "f_operacion"):
        date_key = "f_valor"

    wb = None
    try:
        wb, sheet = _open_sheet(path, sheet_index)
        book = wb if hasattr(wb, "datemode") else None

        header_row, colmap = _find_header_row(sheet)
        start_row = header_row + 1

        if "f_operacion" in colmap or "f_valor" in colmap:
            return _parse_wide_rows(book, sheet, colmap, start_row, source_id, date_key)
        if "movimiento" in colmap and "importe" in colmap:
            return _parse_compact_rows(book, sheet, colmap, start_row, source_id)
        raise ValueError("Unsupported column layout: %s" % list(colmap.keys()))
    finally:
        if wb is not None and path.suffix.lower() == ".xlsx":
            wb.close()
